'''
Created on Jan. 21, 2022

@author: amitmittal
@description: MultipleSubtask class creates multiple subtasks in Jira using python
It accepts the variable from the run file
'''

import requests
from requests.auth import HTTPBasicAuth
from automation import key
import json
import openpyxl

 
class MultipleSubtask():
    
    url = "https://freetestingapi.atlassian.net/rest/api/2/issue"
    auth = HTTPBasicAuth("hi.amitmittal@gmail.com", key.Key().key)
    headers = {
    "Accept": "application/json",
    "Content-Type": "application/json"
    }
    
    def __init__(self, issueType, parent_key, project_key, pro_type):
        
        self.issueType = issueType
        self.parent_key = parent_key
        self.project_key = project_key
        self.pro_type = pro_type
    
    '''
    @description: rollOut method creates multiple subtasks as per flow diagram of the roll out process available in rollout sheet (in Issue.xlsx) 
    '''
    
    def rollOut(self):
    #Access Excel
        wbook = openpyxl.load_workbook('Issue.xlsx') #Access the file
        shee1 = wbook['rollout'] #Access Sheet1
        rows = shee1.max_row #Number of rows
        
        for j in range (1, rows):
            label = (shee1.cell(j+1,4).value).split()
            payload = json.dumps( {
                
                "fields": {
                    "summary": shee1.cell(j+1,1).value,
                    "description": shee1.cell(j+1,2).value,
                    "issuetype": {
                        "id": self.issueType,
                        },
                    "parent":{
                        "key": self.parent_key
                        },
                    "project":{
                        "key": self.project_key
                        },
                    "assignee": {
                    "id": shee1.cell(j+1,3).value,
                    },
                    "labels":label           
                    }
                }
            )

            response = requests.request(
               "POST",
               self.url,
               data=payload,
               headers=self.headers,
               auth=self.auth,
               )
            
            #print(response.text)
            print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))

    '''
    @description: rapidStart method creates multiple subtasks as per flow diagram of the rapid start process available in rapidstart sheet (in Issue.xlsx) 
    '''
                 
    def rapidStart(self):
    #Access Excel
        wbook = openpyxl.load_workbook('Issue.xlsx') #Access the file
        shee1 = wbook['rapidstart'] #Access Sheet1
        rows = shee1.max_row #Number of rows
        
        for j in range (1, rows):
            label = (shee1.cell(j+1,4).value).split()
            payload = json.dumps( {
                
                "fields": {
                    "summary": shee1.cell(j+1,1).value,
                    "description": shee1.cell(j+1,2).value,
                    "issuetype": {
                        "id": self.issueType,
                        },
                    "parent":{
                        "key": self.parent_key
                        },
                    "project":{
                        "key": self.project_key
                        },
                    "assignee": {
                    "id": shee1.cell(j+1,3).value,
                    },
                    "labels":label           
                    }
                }
            )

            response = requests.request(
               "POST",
               self.url,
               data=payload,
               headers=self.headers,
               auth=self.auth,
               )
            
            #print(response.text)
            print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))

    '''
    @description: rapidResponse method creates multiple subtasks as per flow diagram of the rapid response process available in rapidresponse sheet (in Issue.xlsx) 
    '''
            
    def rapidResponse(self):
    #Access Excel
        wbook = openpyxl.load_workbook('Issue.xlsx') #Access the file
        shee1 = wbook['rapidresponse'] #Access Sheet1
        rows = shee1.max_row #Number of rows
        
        for j in range (1, rows):
            label = (shee1.cell(j+1,4).value).split()
            payload = json.dumps( {
                
                "fields": {
                    "summary": shee1.cell(j+1,1).value,
                    "description": shee1.cell(j+1,2).value,
                    "issuetype": {
                        "id": self.issueType,
                        },
                    "parent":{
                        "key": self.parent_key
                        },
                    "project":{
                        "key": self.project_key
                        },
                    "assignee": {
                    "id": shee1.cell(j+1,3).value,
                    },
                    "labels":label           
                    }
                }
            )

            response = requests.request(
               "POST",
               self.url,
               data=payload,
               headers=self.headers,
               auth=self.auth,
               )
            
            #print(response.text)
            print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))

    '''
    @description: multiSubtask method asks user if it is Initial deployment or Rollout and accordingly calls the method as per the flow (rollOut/rapidStart/rapidResponse) 
    It is solely to create multiple subtasks. 
    '''
    
    def multiSubtask(self, dep_value):
        print(self.pro_type)
        
        if (dep_value == "Initial Deployment"):
            
            if (self.pro_type == "RapidStart"):
                self.rapidStart()
            
            else:
                self.rapidResponse()
                
        else:
            self.rollOut()

    
