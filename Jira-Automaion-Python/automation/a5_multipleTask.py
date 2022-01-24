'''
Created on Jan. 21, 2022

@author: amitmittal
'''

import requests
from requests.auth import HTTPBasicAuth
import json
import openpyxl


#1. Get the project key from the dashboard and go to the link: https://freetestingapi.atlassian.net/rest/api/latest/issue/createmeta?projecyKeys="Enter here"
#2. Get the key for the Task
 
class MultipleTask():
    
    url = "https://freetestingapi.atlassian.net/rest/api/2/issue"
    auth = HTTPBasicAuth("hi.amitmittal@gmail.com", "4x7NQ3XzzcKZog9t6HsXAFBC")
    headers = {
    "Accept": "application/json",
    "Content-Type": "application/json"
    }

    def __init__(self, issueType, project_key):

        self.issueType = issueType
        self.project_key = project_key
    
    def multiTask(self):
        
        #Access Excel
        wb = openpyxl.load_workbook('Issue.xlsx') #Access the file
        sh1 = wb['Sheet1'] #Access Sheet1
        row = sh1.max_row #Number of rows

        for i in range (1, row):
            payload = json.dumps( {
                
                "fields": {
                    "summary": sh1.cell(i+1,1).value,
                    "description": sh1.cell(i+1,2).value,
                    "issuetype": {
                        "id": self.issueType,
                        },
                    "project":{
                        "key": self.project_key
                        }
                    }
                }
            )
        
            response = requests.request(
               "POST",
               self.url,
               data=payload,
               headers=self.headers,
               auth=self.auth,
               )
            
            #print(response.text)
            print(json.dumps(json.loads(response.text), sort_keys=True, indent=4, separators=(",", ": ")))
            
